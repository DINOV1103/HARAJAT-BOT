"""
Moliya Menejeri - Asosiy mantiqiy kod.
Ishga tushirish: python bot.py
"""

import asyncio
import logging
import os
import re
from datetime import date, timedelta

from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from aiogram import Bot, Dispatcher, Router, F
from aiogram.client.default import DefaultBotProperties
from aiogram.filters import CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, FSInputFile

from database import (
    init_db, add_user, get_users_count, get_all_user_ids,
    add_transaction, get_transactions_by_date, get_transactions_by_range, get_all_transactions_for_excel,
    add_debt, get_active_debts, mark_debt_as_paid,
    add_note, get_notes, delete_note
)

from keyboards import (
    main_menu_keyboard, cancel_keyboard, stats_keyboard,
    debts_keyboard, notes_keyboard, admin_keyboard, debt_type_inline, calendar_keyboard, category_inline_keyboard,
    BTN_EXPENSE, BTN_INCOME, BTN_STATS, BTN_DEBTS, BTN_NOTES, BTN_CALENDAR, BTN_ADMIN, BTN_MAIN_MENU,
    BTN_STATS_TODAY, BTN_STATS_WEEK, BTN_STATS_MONTH, BTN_STATS_TOTAL, BTN_EXPORT_EXCEL,
    BTN_DEBTS_ADD, BTN_DEBTS_LIST, BTN_NOTES_ADD, BTN_NOTES_LIST,
    BTN_ADMIN_USERS, BTN_ADMIN_BROADCAST, BTN_ADMIN_BACKUP
)

load_dotenv()
BOT_TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = 1691140865  # Admin ID[span_0](start_span)[span_0](end_span)

logging.basicConfig(level=logging.INFO)
router = Router()


class Form(StatesGroup):
    waiting_expense = State()
    waiting_expense_cat = State()
    waiting_income = State()
    waiting_income_cat = State()
    waiting_debt_input = State()
    waiting_note_input = State()
    waiting_admin_broadcast = State()


# ---------- Yordamchi Funksiyalar ----------

def format_amount(amount: float) -> str:
    return f"{int(amount):,}".replace(",", " ") + " so'm"


def format_date_human(date_str: str) -> str:
    y, m, d = date_str.split("-")
    return f"{d}.{m}.{y}"


def parse_amount_and_description(text: str):
    text = text.strip()
    match = re.match(r"^([\d\s.,]+)\s*(.*)$", text)
    if not match:
        return None, None
    raw_number = match.group(1)
    description = match.group(2).strip() or "Izohsiz"
    clean_number = re.sub(r"[\s.,]", "", raw_number)
    if not clean_number.isdigit() or clean_number == "":
        return None, None
    amount = float(clean_number)
    return (amount, description) if amount > 0 else (None, None)


def parse_debt_text(text: str):
    text = text.strip()
    match = re.match(r"^(.+?)\s+([\d\s.,]+)$", text)
    if not match:
        return None, None
    name = match.group(1).strip()
    raw_amount = match.group(2)
    clean_amount = re.sub(r"[\s.,]", "", raw_amount)
    if not clean_amount.isdigit():
        return None, None
    return name, float(clean_amount)


def format_comprehensive_summary(rows, title: str) -> str:
    if not rows:
        return f"<b>{title}</b>\n\nBu davrda hech qanday amallar bajarilmadi."

    expenses = [r for r in rows if r[-1] == 'expense']
    incomes = [r for r in rows if r[-1] == 'income']
    total_exp = sum(r[0] for r in expenses)
    total_inc = sum(r[0] for r in incomes)

    lines = [f"<b>{title}</b>", ""]

    # Toifalar bo'yicha guruhlash statistikasi
    cat_summary = {}
    
    if len(rows[0]) == 5:  # Rangeli hisobot (Hafta/Oy)
        for amount, cat, desc, txn_date, t_type in rows:
            cat_summary[cat] = cat_summary.get(cat, 0) + amount
    else:  # Kunlik oddiy hisobot
        for amount, cat, desc, t_type in rows:
            cat_summary[cat] = cat_summary.get(cat, 0) + amount

    lines.append("📌 <b>Toifalar kesimida taqsimot:</b>")
    for cat, amt in cat_summary.items():
        lines.append(f"• {cat}: <b>{format_amount(amt)}</b>")
        
    lines.append("\n📝 <b>Batafsil amallar ro'yxati:</b>")
    if len(rows[0]) == 5:
        grouped = {}
        for amount, cat, desc, txn_date, t_type in rows:
            grouped.setdefault(txn_date, []).append((amount, cat, desc, t_type))
        for d in sorted(grouped.keys()):
            lines.append(f"📅 <b>{format_date_human(d)}:</b>")
            for amount, cat, desc, t_type in grouped[d]:
                icon = "💸" if t_type == "expense" else "💰"
                lines.append(f"   • {icon} {format_amount(amount)} — {cat} ({desc})")
    else:
        for amount, cat, desc, t_type in rows:
            icon = "💸" if t_type == "expense" else "💰"
            lines.append(f"• {icon} {format_amount(amount)} — {cat} ({desc})")

    lines.append("\n" + "—" * 20)
    lines.append(f"💰 Jami daromad: <b>{format_amount(total_inc)}</b>")
    lines.append(f"💸 Jami harajat: <b>{format_amount(total_exp)}</b>")
    lines.append(f"⚖️ Sof foyda/Zarar: <b>{format_amount(total_inc - total_exp)}</b>")
    return "\n".join(lines)


# ---------- Tizim va Navigatsiya Handlerlari ----------

@router.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext):
    await state.clear()
    add_user(message.from_user.id, message.from_user.username, message.from_user.full_name)
    is_admin = (message.from_user.id == ADMIN_ID)
    await message.answer(
        f"Salom, <b>{message.from_user.full_name}</b>! Professional Moliyaviy Menejer botiga xush kelibsiz.\n\n"
        "O'z mablag'laringizni toifalar yordamida nazorat qiling va Excel hisobotlarni yuklab oling.",
        reply_markup=main_menu_keyboard(is_admin)
    )


@router.message(F.text == BTN_MAIN_MENU)
async def back_to_menu(message: Message, state: FSMContext):
    await state.clear()
    is_admin = (message.from_user.id == ADMIN_ID)
    await message.answer("🏠 Asosiy menyu.", reply_markup=main_menu_keyboard(is_admin))


# ---------- Harajat Bo'limi (Kategoriyali) ----------

@router.message(F.text == BTN_EXPENSE)
async def ask_expense(message: Message, state: FSMContext):
    await state.set_state(Form.waiting_expense)
    await message.answer("✏️ Harajat summasi va sababini kiriting.\nMasalan: <i>55000 tushlik do'stlar bilan</i>", reply_markup=cancel_keyboard())


@router.message(Form.waiting_expense)
async def process_expense_amt(message: Message, state: FSMContext):
    if message.text == BTN_MAIN_MENU:
        await back_to_menu(message, state)
        return
    amount, description = parse_amount_and_description(message.text or "")
    if amount is None:
        await message.answer("⚠️ Xatolik. Summani raqam bilan boshlang.\nMasalan: <i>15000 avtobus</i>")
        return
    await state.update_data(amt=amount, desc=description, type_="expense")
    await state.set_state(Form.waiting_expense_cat)
    await message.answer("📁 Ushbu harajat qaysi <b>toifaga (kategoriyaga)</b> tegishli?", reply_markup=category_inline_keyboard("expense"))


# ---------- Daromad Bo'limi (Kategoriyali) ----------

@router.message(F.text == BTN_INCOME)
async def ask_income(message: Message, state: FSMContext):
    await state.set_state(Form.waiting_income)
    await message.answer("✏️ Kelgan daromad summasi va manbasini kiriting.\nMasalan: <i>4500000 bonus oylikdan</i>", reply_markup=cancel_keyboard())


@router.message(Form.waiting_income)
async def process_income_amt(message: Message, state: FSMContext):
    if message.text == BTN_MAIN_MENU:
        await back_to_menu(message, state)
        return
    amount, description = parse_amount_and_description(message.text or "")
    if amount is None:
        await message.answer("⚠️ Xatolik. Summani raqam bilan boshlang.\nMasalan: <i>500000 kassa</i>")
        return
    await state.update_data(amt=amount, desc=description, type_="income")
    await state.set_state(Form.waiting_income_cat)
    await message.answer("📁 Ushbu daromad qaysi <b>toifaga (kategoriyaga)</b> tegishli?", reply_markup=category_inline_keyboard("income"))


# ---------- Toifani Saqlash (Callback Query) ----------

@router.callback_query(F.data.startswith("setcat_"))
async def save_transaction_with_cat(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    amount = data.get("amt")
    description = data.get("desc")
    t_type = data.get("type")
    
    if amount is None or t_type is None:
        await callback.answer("Ma'lumot topilmadi. Qayta urinib ko'ring.", show_alert=True)
        return
        
    category = callback.data.split("_")[1]
    today = date.today().isoformat()
    
    add_transaction(callback.from_user.id, t_type, amount, category, description, today)
    await state.clear()
    await callback.message.delete()
    
    is_admin = (callback.from_user.id == ADMIN_ID)
    status_text = "💸 Harajat saqlandi" if t_type == "expense" else "💰 Daromad saqlandi"
    await callback.message.answer(
        f"✅ <b>{status_text}!</b>\n\n💰 Summa: <b>{format_amount(amount)}</b>\n📁 Toifa: {category}\n📝 Izoh: {description}", 
        reply_markup=main_menu_keyboard(is_admin)
    )
    await callback.answer()


@router.callback_query(F.data == "cat_cancel")
async def cancel_cat_cb(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.delete()
    is_admin = (callback.from_user.id == ADMIN_ID)
    await callback.message.answer("❌ Amaliyot bekor qilindi.", reply_markup=main_menu_keyboard(is_admin))
    await callback.answer()


# ---------- Statistika va Excel Eksport ----------

@router.message(F.text == BTN_STATS)
async def show_stats_menu(message: Message):
    await message.answer("📊 Hisobotlar bo'limi. Kerakli davrni tanlang yoki ma'lumotlarni Excel formatida yuklab oling:", reply_markup=stats_keyboard())


@router.message(F.text == BTN_STATS_TODAY)
async def today_summary(message: Message):
    today = date.today().isoformat()
    rows = get_transactions_by_date(message.from_user.id, today)
    await message.answer(format_comprehensive_summary(rows, f"📋 Bugungi amallar ({format_date_human(today)})"))


@router.message(F.text == BTN_STATS_WEEK)
async def week_summary(message: Message):
    today = date.today()
    start = today - timedelta(days=today.weekday())
    rows = get_transactions_by_range(message.from_user.id, start.isoformat(), today.isoformat())
    await message.answer(format_comprehensive_summary(rows, "📆 Haftalik jami ko'rsatkichlar"))


@router.message(F.text == BTN_STATS_MONTH)
async def month_summary(message: Message):
    today = date.today()
    start = today.replace(day=1)
    rows = get_transactions_by_range(message.from_user.id, start.isoformat(), today.isoformat())
    await message.answer(format_comprehensive_summary(rows, "🗓 Joriy oylik ko'rsatkichlar"))


@router.message(F.text == BTN_STATS_TOTAL)
async def total_summary(message: Message):
    rows = get_transactions_by_range(message.from_user.id, "1970-01-01", date.today().isoformat())
    await message.answer(format_comprehensive_summary(rows, "📈 Tizimdagi umumiy moliyaviy balans"))


@router.message(F.text == BTN_EXPORT_EXCEL)
async def export_to_excel_handler(message: Message):
    rows = get_all_transactions_for_excel(message.from_user.id)
    if not rows:
        await message.answer("📥 Bazangiz bo'sh, excel fayl yaratish uchun hech qanday ma'lumot yo'q.")
        return
        
    await message.answer("⏳ Excel hisobotingiz shakllantirilmoqda, iltimos kuting...")
    
    # Excel hujjati yaratish (openpyxl)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Moliyaviy Hisobot"
    
    # Dizayn va stillar
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    headers = ["Sana", "Tur", "Toifa (Kategoriya)", "Summa", "Izoh / Sabab"]
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    for row_idx, data in enumerate(rows, 2):
        sana, tur, toifa, summa, izoh = data
        tur_uz = "Daromad 💰" if tur == "income" else "Harajat 💸"
        ws.append([format_date_human(sana), tur_uz, toifa, summa, izoh])
        
        for col_num in range(1, 6):
            c = ws.cell(row=row_idx, column=col_num)
            c.font = data_font
            if col_num in [1, 2]:
                c.alignment = align_center
            elif col_num == 4:
                c.number_format = '#,##0" so\'m"'
                c.alignment = align_left
            else:
                c.alignment = align_left

    # Ustun kengliklarini avtomatlashtirish
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    file_path = f"Hisobot_{message.from_user.id}.xlsx"
    wb.save(file_path)
    
    # Telegramga yuborish
    excel_file = FSInputFile(file_path, filename="Mening_Moliyaviy_Hisobotim.xlsx")
    await message.answer_document(document=excel_file, caption="✅ Sizning barcha davrlardagi batafsil moliyaviy hisobotingiz tayyor bo'ldi!")
    
    if os.path.exists(file_path):
        os.remove(file_path)


# ---------- Qarz Daftari Bo'limi ----------

@router.message(F.text == BTN_DEBTS)
async def show_debts_menu(message: Message):
    await message.answer("📒 Qarz daftari bo'limi:", reply_markup=debts_keyboard())


@router.message(F.text == BTN_DEBTS_ADD)
async def ask_debt(message: Message, state: FSMContext):
    await state.set_state(Form.waiting_debt_input)
    await message.answer("✏️ Kim bilan qarz munosabati bo'ldi va qancha?\nMasalan: <i>Jasur 500000</i>", reply_markup=cancel_keyboard())


@router.message(Form.waiting_debt_input)
async def process_debt_input(message: Message, state: FSMContext):
    if message.text == BTN_MAIN_MENU:
        await back_to_menu(message, state)
        return
    name, amount = parse_debt_text(message.text or "")
    if not name or amount is None:
        await message.answer("⚠️ Xato shakl. Ism va summani bo'sh joy qoldirib yozing.\nMasalan: <i>Asror 250000</i>")
        return
    await state.update_data(d_name=name, d_amount=amount)
    await message.answer(f"👤 {name} — {format_amount(amount)}\n\nUshbu qarz turini belgilang:", reply_markup=debt_type_inline())


@router.callback_query(F.data.startswith("debt_type_"))
async def save_debt_callback(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    name = data.get("d_name")
    amount = data.get("d_amount")
    if not name or not amount:
        await callback.answer("Xatolik ro'y berdi.", show_alert=True)
        return
    g_type = callback.data.split("_")[2]
    add_debt(callback.from_user.id, name, amount, g_type)
    await state.clear()
    await callback.message.delete()
    is_admin = (callback.from_user.id == ADMIN_ID)
    type_str = "Qarz berildi 💸" if g_type == "lent" else "Qarz olindi 💰"
    await callback.message.answer(f"✅ Qarz saqlandi:\n👤 {name} — <b>{format_amount(amount)}</b> ({type_str})", reply_markup=main_menu_keyboard(is_admin))
    await callback.answer()


@router.callback_query(F.data == "debt_cancel")
async def cancel_debt_cb(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.delete()
    is_admin = (callback.from_user.id == ADMIN_ID)
    await callback.message.answer("❌ Qarz yozish bekor qilindi.", reply_markup=main_menu_keyboard(is_admin))
    await callback.answer()


@router.message(F.text == BTN_DEBTS_LIST)
async def list_debts(message: Message):
    rows = get_active_debts(message.from_user.id)
    if not rows:
        await message.answer("📒 Hozirda faol qarzlaringiz mavjud emas.")
        return
    await message.answer("📜 <b>Sizning faol qarzlaringiz ro'yxati:</b>")
    for d_id, name, amount, d_type in rows:
        direction = "Bergan qarzim (menga qaytadi) ➡️" if d_type == "lent" else "Olgan qarzim (qaytarishim kerak) ⬅️"
        txt = f"👤 <b>{name}</b>\n💵 Summa: {format_amount(amount)}\n📋 Holat: {direction}"
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ To'landi (Yopish)", callback_data=f"pay_{d_id}")]
        ])
        await message.answer(txt, reply_markup=kb)


@router.callback_query(F.data.startswith("pay_"))
async def close_debt(callback: CallbackQuery):
    debt_id = int(callback.data.split("_")[1])
    mark_debt_as_paid(debt_id, callback.from_user.id)
    await callback.message.edit_text(callback.message.text + "\n\n✅ <b>Ushbu qarz to'liq yopildi va arxivlandi!</b>")
    await callback.answer("Qarz yopildi")


# ---------- Eslatmalar Bo'limi ----------

@router.message(F.text == BTN_NOTES)
async def show_notes_menu(message: Message):
    await message.answer("📝 Eslatmalar bo'limi:", reply_markup=notes_keyboard())


@router.message(F.text == BTN_NOTES_ADD)
async def ask_note(message: Message, state: FSMContext):
    await state.set_state(Form.waiting_note_input)
    await message.answer("✏️ Eslatmani yozib yuboring:", reply_markup=cancel_keyboard())


@router.message(Form.waiting_note_input)
async def save_note_msg(message: Message, state: FSMContext):
    if message.text == BTN_MAIN_MENU:
        await back_to_menu(message, state)
        return
    if not message.text:
        await message.answer("⚠️ Matnli eslatma kiriting.")
        return
    add_note(message.from_user.id, message.text)
    await state.clear()
    is_admin = (message.from_user.id == ADMIN_ID)
    await message.answer("✅ Eslatma muvaffaqiyatli saqlandi!", reply_markup=main_menu_keyboard(is_admin))


@router.message(F.text == BTN_NOTES_LIST)
async def list_notes(message: Message):
    rows = get_notes(message.from_user.id)
    if not rows:
        await message.answer("📝 Eslatmalaringiz ro'yxati bo'sh.")
        return
    await message.answer("📌 <b>Sizning eslatmalaringiz:</b>")
    for n_id, content in rows:
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"del_note_{n_id}")]
        ])
        await message.answer(f"▫️ {content}", reply_markup=kb)


@router.callback_query(F.data.startswith("del_note_"))
async def remove_note_cb(callback: CallbackQuery):
    note_id = int(callback.data.split("_")[2])
    delete_note(note_id, callback.from_user.id)
    await callback.message.delete()
    await callback.answer("Eslatma o'chirildi")


# ---------- Kalendar Bo'limi ----------

@router.message(F.text == BTN_CALENDAR)
async def show_calendar(message: Message):
    today = date.today()
    await message.answer("📅 Tarixni ko'rish uchun kunni tanlang:", reply_markup=calendar_keyboard(today.year, today.month))


@router.callback_query(F.data == "cal_ignore")
async def cal_ignore(callback: CallbackQuery):
    await callback.answer()


@router.callback_query(F.data.startswith("cal_nav_"))
async def cal_navigate(callback: CallbackQuery):
    _, _, year, month = callback.data.split("_")
    await callback.message.edit_reply_markup(reply_markup=calendar_keyboard(int(year), int(month)))
    await callback.answer()


@router.callback_query(F.data.startswith("cal_day_"))
async def cal_day_selected(callback: CallbackQuery):
    _, _, year, month, day = callback.data.split("_")
    selected_date = date(int(year), int(month), int(day)).isoformat()
    rows = get_transactions_by_date(callback.from_user.id, selected_date)
    text = format_comprehensive_summary(rows, f"📅 {format_date_human(selected_date)} kunidagi hisobot")
    await callback.message.answer(text)
    await callback.answer()


# ---------- Admin Panel Bo'limi ----------

@router.message(F.text == BTN_ADMIN)
async def show_admin_panel(message: Message):
    if message.from_user.id != ADMIN_ID:
        return
    await message.answer("🛡 Admin panel boshqaruv tizimi:", reply_markup=admin_keyboard())


@router.message(F.text == BTN_ADMIN_USERS)
async def admin_users_count(message: Message):
    if message.from_user.id != ADMIN_ID:
        return
    count = get_users_count()
    await message.answer(f"👥 Bot ichidagi jami a'zolar soni: <b>{count} ta</b>")


@router.message(F.text == BTN_ADMIN_BACKUP)
async def admin_backup_database(message: Message):
    if message.from_user.id != ADMIN_ID:
        return
    if os.path.exists(DB_NAME):
        db_file = FSInputFile(DB_NAME, filename=f"Backup_{date.today().isoformat()}.db")
        await message.answer_document(document=db_file, caption="💾 Ma'lumotlar bazasining joriy zaxira nusxasi (Backup) tayyor.")
    else:
        await message.answer("❌ Ma'lumotlar bazasi fayli topilmadi.")


@router.message(F.text == BTN_ADMIN_BROADCAST)
async def admin_broadcast_ask(message: Message, state: FSMContext):
    if message.from_user.id != ADMIN_ID:
        return
    await state.set_state(Form.waiting_admin_broadcast)
    await message.answer("📢 Barcha foydalanuvchilarga yuboriladigan xabar matnini kiriting:", reply_markup=cancel_keyboard())


@router.message(Form.waiting_admin_broadcast)
async def admin_broadcast_send(message: Message, state: FSMContext, bot: Bot):
    if message.from_user.id != ADMIN_ID:
        return
    if message.text == BTN_MAIN_MENU:
        await back_to_menu(message, state)
        return

    u_ids = get_all_user_ids()
    sent_count = 0
    await message.answer("⏳ Tarqatish boshlandi, iltimos kuting...")

    for uid in u_ids:
        try:
            await bot.send_message(chat_id=uid, text=message.text, parse_mode="HTML")
            sent_count += 1
            await asyncio.sleep(0.05)
        except Exception:
            pass

    await state.clear()
    await message.answer(f"📢 Tarqatish yakunlandi.\n✅ {sent_count} ta a'zoga muvaffaqiyatli yetib bordi.", reply_markup=main_menu_keyboard(True))


# ---------- Loyihani Ishga Tushirish ----------

async def main():
    if not BOT_TOKEN:
        raise RuntimeError("BOT_TOKEN topilmadi! .env faylini to'g'ri sozlang.")
    init_db()
    bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode="HTML"))
    dp = Dispatcher()
    dp.include_router(router)
    logging.info("Bot professional rejimda muvaffaqiyatli ishga tushdi!")
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
    
